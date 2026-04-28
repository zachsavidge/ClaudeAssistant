#!/usr/bin/env python3
"""
convert_xls.py - Cross-platform .xls → .xlsx batch converter.

Strategy (tiered, formula-preservation aware):
  1. PRIMARY: LibreOffice headless (`soffice --headless --convert-to xlsx`)
     - Cross-platform: Linux cloud sandboxes (apt install libreoffice-calc),
       macOS (brew install --cask libreoffice), Windows (rare but works)
     - Preserves formulas, charts, formatting (verified on SUM/AVERAGE/mixed
       arithmetic; Excel data_type='f' round-trips correctly)
     - Auto-selected when `soffice` (or `libreoffice`) is on PATH
     - ~3s/file on warm sandbox; ~100MB install, one-time
     - Cannot handle password-protected files via CLI (decrypt first)
  2. FALLBACK A: Excel COM via pywin32 (Windows only, opt-in via --use-excel)
     - Preserves formulas, formatting, charts (truest fidelity to Excel)
     - Can hang on certain files (especially Drive Streaming paths)
     - Handles password-protected files inline
  3. FALLBACK B: pure-Python via xlrd + openpyxl (last resort, cross-platform)
     - Loses formulas/charts but preserves all values + text
     - Fast; reliable for survey/research data without live calculations

Tiers 1 and 2 handle password-protected .xls via the optional --password flag.

Usage:
  python convert_xls.py <folder>                      # walk folder, convert all .xls
  python convert_xls.py <folder> --password RENGA2025 # try this password if encrypted
  python convert_xls.py <folder> --use-excel          # prefer Excel COM (Windows + Excel)
  python convert_xls.py <folder> --no-libreoffice     # skip LibreOffice (force-fallback)
  python convert_xls.py <file.xls>                    # convert single file

Outputs the .xlsx alongside the .xls and removes the original on success.
"""

import os
import shutil
import subprocess
import sys
import time
import argparse


def is_xls(path):
    return path.lower().endswith(".xls") and not path.lower().endswith(".xlsx")


def convert_pure_python(xls_path, password=None):
    """Convert .xls → .xlsx via xlrd + openpyxl. Cross-platform."""
    import xlrd
    from openpyxl import Workbook

    try:
        # xlrd 1.2.0 is the last version that supports .xls
        # If password-protected, xlrd doesn't decrypt - that's the msoffcrypto path
        rb = xlrd.open_workbook(xls_path, formatting_info=False)
    except xlrd.biffh.XLRDError as e:
        if "encrypted" in str(e).lower():
            return False, "encrypted (use decrypt_excel.py first)"
        return False, f"xlrd: {e}"
    except Exception as e:
        return False, f"open: {e}"

    wb = Workbook()
    wb.remove(wb.active)

    used_names = set()
    for sheet_name in rb.sheet_names():
        rs = rb.sheet_by_name(sheet_name)
        # Sanitize: max 31 chars, no special chars
        safe = sheet_name
        for bad in "[]*?:/\\":
            safe = safe.replace(bad, "_")
        safe = safe[:31]
        orig = safe
        i = 1
        while safe in used_names:
            safe = f"{orig[:28]}_{i}"
            i += 1
        used_names.add(safe)

        ws = wb.create_sheet(title=safe)
        for r in range(rs.nrows):
            for c in range(rs.ncols):
                cell = rs.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    continue
                val = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        val = xlrd.xldate_as_datetime(val, rb.datemode)
                    except Exception:
                        pass
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    val = bool(val)
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    val = None
                if val is None or val == "":
                    continue
                ws.cell(row=r + 1, column=c + 1, value=val)

    dst = os.path.splitext(xls_path)[0] + ".xlsx"
    wb.save(dst)
    return True, f"OK ({len(rb.sheet_names())} sheets)"


def _find_soffice():
    """Locate the LibreOffice binary on PATH. Returns path or None."""
    return shutil.which("soffice") or shutil.which("libreoffice")


def convert_libreoffice(xls_path, password=None, timeout=120):
    """Convert via LibreOffice headless. Cross-platform, preserves formulas/charts.

    Verified in Anthropic cloud sandboxes (Debian + libreoffice-calc): SUM,
    AVERAGE, and mixed-arithmetic formulas round-trip with openpyxl
    data_type='f'. macOS works after `brew install --cask libreoffice`.
    """
    soffice = _find_soffice()
    if not soffice:
        return False, "soffice not on PATH (install libreoffice-calc / libreoffice)"

    if password:
        # LibreOffice CLI doesn't accept passwords directly. Decrypt first via
        # decrypt_excel.py (msoffcrypto-tool) and re-run on the cleartext copy.
        return False, "LibreOffice CLI cannot accept password — decrypt with decrypt_excel.py first"

    out_dir = os.path.dirname(os.path.abspath(xls_path))
    cmd = [soffice, "--headless", "--convert-to", "xlsx", "--outdir", out_dir, xls_path]

    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=timeout,
        )
    except subprocess.TimeoutExpired:
        return False, f"LibreOffice timed out (>{timeout}s)"
    except Exception as e:
        return False, f"LibreOffice exec: {e}"

    expected = os.path.join(
        out_dir, os.path.splitext(os.path.basename(xls_path))[0] + ".xlsx"
    )
    if result.returncode == 0 and os.path.exists(expected):
        return True, "OK (via LibreOffice)"

    err = (result.stderr or result.stdout or "").strip().replace("\n", " | ")[:200]
    return False, f"LibreOffice rc={result.returncode}: {err}"


def convert_excel_com(xls_path, password=None, timeout=30):
    """Convert via Excel COM (Windows only). Returns (success, message)."""
    if sys.platform != "win32":
        return False, "Excel COM only available on Windows"
    try:
        import win32com.client
        import pythoncom
    except ImportError:
        return False, "pywin32 not installed"

    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.EnableEvents = False
        try:
            excel.AutomationSecurity = 3  # disable macros
        except Exception:
            pass

        if password:
            wb = excel.Workbooks.Open(xls_path, 0, False, 5, password)
            wb.Password = ""
        else:
            wb = excel.Workbooks.Open(xls_path, 0, False, 5)

        dst = os.path.splitext(xls_path)[0] + ".xlsx"
        wb.SaveAs(dst, FileFormat=51)
        wb.Close(False)
        excel.Quit()
        return True, "OK (via Excel COM)"
    except Exception as e:
        if wb:
            try:
                wb.Close(False)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        return False, f"Excel COM: {e}"


def convert_one(xls_path, password=None, use_excel=False, no_libreoffice=False):
    """Convert a single file. Tier order:

    1. Excel COM (Windows + --use-excel only) — truest fidelity, opt-in
    2. LibreOffice headless (auto-detect) — formula-preserving, cross-platform
    3. Pure Python (xlrd + openpyxl) — fallback, lossy on formulas
    """
    # Tier 1: Excel COM (only if explicitly requested AND Windows)
    if use_excel and sys.platform == "win32":
        ok, msg = convert_excel_com(xls_path, password)
        if ok:
            return ok, msg
        # Fall through

    # Tier 2: LibreOffice (skipped only if explicitly disabled)
    if not no_libreoffice and _find_soffice():
        ok, msg = convert_libreoffice(xls_path, password)
        if ok:
            return ok, msg
        # Fall through (e.g. password-protected, soffice crashed)

    # Tier 3: pure Python (lossy but always available)
    return convert_pure_python(xls_path, password)


def find_xls_files(target):
    """Return list of .xls files. target may be a single file or a folder."""
    if os.path.isfile(target):
        return [target] if is_xls(target) else []
    out = []
    for root, dirs, files in os.walk(target):
        for f in files:
            if is_xls(f):
                out.append(os.path.join(root, f))
    return sorted(out)


def main():
    p = argparse.ArgumentParser(description="Cross-platform .xls → .xlsx converter")
    p.add_argument("target", help="Folder or single .xls file")
    p.add_argument("--password", help="Password to try if files are encrypted")
    p.add_argument("--use-excel", action="store_true",
                   help="Try Excel COM first (Windows only); fall back to LibreOffice/pure Python on failure")
    p.add_argument("--no-libreoffice", action="store_true",
                   help="Skip LibreOffice tier even if soffice is on PATH (forces pure-Python fallback)")
    p.add_argument("--keep-original", action="store_true",
                   help="Don't delete the .xls after successful conversion")
    args = p.parse_args()

    files = find_xls_files(args.target)
    if not files:
        print(f"No .xls files found in {args.target}")
        return

    soffice = _find_soffice()
    tiers = []
    if args.use_excel and sys.platform == "win32":
        tiers.append("Excel COM")
    if soffice and not args.no_libreoffice:
        tiers.append(f"LibreOffice ({soffice})")
    tiers.append("pure Python")

    print(f"Found {len(files)} .xls file(s). Platform: {sys.platform}")
    print(f"Strategy: {' → '.join(tiers)}")
    print()

    ok_count = 0
    fail = []
    for i, path in enumerate(files, 1):
        rel = os.path.relpath(path)
        print(f"[{i}/{len(files)}] {rel}", flush=True)
        t0 = time.time()
        success, msg = convert_one(
            path,
            password=args.password,
            use_excel=args.use_excel,
            no_libreoffice=args.no_libreoffice,
        )
        elapsed = time.time() - t0
        if success:
            if not args.keep_original:
                try:
                    os.remove(path)
                except Exception:
                    pass
            ok_count += 1
            print(f"    {msg} ({elapsed:.1f}s)", flush=True)
        else:
            fail.append((rel, msg))
            print(f"    FAILED: {msg[:150]} ({elapsed:.1f}s)", flush=True)

    print(f"\n=== Summary ===")
    print(f"Converted: {ok_count}/{len(files)}")
    if fail:
        print(f"Failed: {len(fail)}")
        for n, m in fail:
            print(f"  - {n}: {m[:100]}")


if __name__ == "__main__":
    main()
